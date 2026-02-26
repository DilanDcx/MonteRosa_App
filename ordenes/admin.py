from django.contrib import admin
from import_export import resources, fields
from import_export.admin import ImportMixin # <-- CAMBIO 1: Solo importamos el Mixin de Importar
from import_export.widgets import ForeignKeyWidget
from datetime import datetime, date
from django.utils.html import format_html
from django.contrib import messages
from django.urls import path, reverse
from django.shortcuts import redirect
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from .models import OrdenTrabajo, Actividad, OrdenBorrador, OrdenPendiente, OrdenHistorial, BitacoraActividad, Evidencia
from django.contrib.auth.models import User, Group 
from django.contrib.auth.admin import UserAdmin
from django.contrib.admin import helpers

# -------------------------
# 1. LÓGICA DE IMPORTACIÓN
# -------------------------
class ActividadResource(resources.ModelResource):
    orden = fields.Field(
        attribute='orden', 
        column_name='Orden',
        widget=ForeignKeyWidget(OrdenTrabajo, field='numero_orden')
    )
    codigo_operacion = fields.Field(attribute='codigo_operacion', column_name='Operación')
    descripcion = fields.Field(attribute='descripcion', column_name='Txt.brv.oper.')
    puesto_trabajo = fields.Field(attribute='puesto_trabajo', column_name='Pto.tbjo.op.')
    
    def __init__(self, user=None, **kwargs):
        super().__init__(**kwargs)
        self.user = user

    class Meta:
        model = Actividad
        import_id_fields = ('codigo_operacion', 'orden') 
        fields = ('orden', 'codigo_operacion', 'descripcion', 'puesto_trabajo')

    def before_import(self, dataset, using_transactions, dry_run, **kwargs):
        new_headers = []
        desc_count = 0
        for header in dataset.headers:
            if header == 'Descripción':
                desc_count += 1
                if desc_count == 1: new_headers.append('desc_equipo_sap')
                elif desc_count == 2: new_headers.append('desc_ubicacion_sap')
                else: new_headers.append(header)
            else:
                new_headers.append(header)
        dataset.headers = new_headers

    def before_import_row(self, row, **kwargs):
        numero_orden_excel = row.get('Orden')
        
        def parse_fecha_inteligente(valor):
            if not valor: return None
            if isinstance(valor, (datetime, date)): return valor
            try: return datetime.strptime(str(valor).strip(), '%d/%m/%Y').date()
            except: 
                try: return datetime.strptime(str(valor).strip().split(' ')[0], '%Y-%m-%d').date()
                except: return None

        fecha_inicio = parse_fecha_inteligente(row.get('Fe.inic.extrema'))
        fecha_fin = parse_fecha_inteligente(row.get('Fe.fin extrema'))

        prioridad_excel = str(row.get('Prioridad', '')).strip()
        prioridad_db = '4' 
        
        if prioridad_excel and prioridad_excel[0] in ['1', '2', '3', '4']:
            prioridad_db = prioridad_excel[0]
        
        OrdenTrabajo.objects.update_or_create(
            numero_orden=numero_orden_excel,
            defaults={
                'descripcion': row.get('Texto breve'),
                'equipo': row.get('Equipo'),
                'descripcion_equipo': row.get('desc_equipo_sap'),
                'ubicacion': row.get('Ubic.técn.'),
                'ubicacion_tecnica': row.get('desc_ubicacion_sap'),
                'inicio_programado': fecha_inicio,
                'fin_programado': fecha_fin,
                'estado': 'BORRADOR',
                'prioridad': prioridad_db,
                'supervisor': self.user
            }
        )

# -----------------------
# PANTALLA 1: BORRADORES 
# -----------------------
class ActividadInline(admin.TabularInline):
    model = Actividad
    fk_name = 'orden'
    extra = 0
    can_delete = False
    fields = ('codigo_operacion', 'descripcion', 'puesto_trabajo')
    readonly_fields = ('codigo_operacion', 'descripcion', 'puesto_trabajo')
    verbose_name = "Operación Detectada"
    verbose_name_plural = "Operaciones (Actividades)"

class EvidenciaInline(admin.StackedInline): 
    model = Evidencia
    fk_name = 'orden'
    extra = 1
    
    verbose_name = "Fotografía de Respaldo"
    verbose_name_plural = "📸 Panel de Evidencias Fotográficas" 
    
    readonly_fields = ('ver_foto', 'fecha_subida')
    fields = ('tipo', 'ver_foto', 'descripcion', 'fecha_subida')

    def ver_foto(self, obj):
        if obj.foto and hasattr(obj.foto, 'url'):
            from django.utils.html import format_html
            return format_html(
                '<a href="{0}" target="_blank">'
                '<img src="{0}" style="height: 250px; border-radius: 8px; border: 2px solid #ddd; box-shadow: 2px 2px 5px rgba(0,0,0,0.1);"/>'
                '</a>', 
                obj.foto.url
            )
        return "Sin imagen"

    ver_foto.short_description = "Vista Previa de la Imagen"

# <-- CAMBIO 2: Usamos ImportMixin y admin.ModelAdmin para ocultar el botón Exportar
@admin.register(OrdenBorrador)
class OrdenBorradorAdmin(ImportMixin, admin.ModelAdmin):
    resource_class = ActividadResource
    
    list_display = (
        'numero_orden', 
        'descripcion', 
        'ver_equipo',
        'prioridad',
        'codigo_trabajador',
        'ver_fechas',
        'botones_accion' 
    )
    list_editable = ('prioridad', 'codigo_trabajador')
    search_fields = ('numero_orden', 'descripcion', 'codigo_trabajador')
    ordering = ('-id',)
    inlines = [ActividadInline, EvidenciaInline]

    class Media:
        css = { 'all': ('css/flotante.css',) }

    def get_import_resource_kwargs(self, request, *args, **kwargs):
        kwargs = super().get_import_resource_kwargs(request, *args, **kwargs)
        kwargs.update({"user": request.user}) 
        return kwargs
    
    @admin.action(description="✅ Aprobar masivamente")
    def aprobar_masivamente(self, request, queryset):
        seleccionados = request.POST.getlist(helpers.ACTION_CHECKBOX_NAME)
        total_forms = int(request.POST.get('form-TOTAL_FORMS', 0))
        
        for i in range(total_forms):
            orden_id = request.POST.get(f'form-{i}-id')
            codigo = request.POST.get(f'form-{i}-codigo_trabajador')
            
            if orden_id in seleccionados and codigo is not None:
                queryset.model.objects.filter(id=orden_id).update(codigo_trabajador=codigo.strip())

        queryset.update(estado='PENDIENTE')
        self.message_user(request, f"¡Éxito! Se aprobaron y enviaron {queryset.count()} órdenes.", messages.SUCCESS)

    actions = ['aprobar_masivamente']

    def get_queryset(self, request):
        return super().get_queryset(request).filter(estado='BORRADOR')

    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'delete_selected' in actions:
            func, name, _ = actions['delete_selected']
            actions['delete_selected'] = (func, name, "🗑️ Eliminar masivamente")
        return actions

    def get_urls(self):
        urls = super().get_urls()
        from django.urls import path
        custom_urls = [
            path('<int:orden_id>/aprobar/', self.admin_site.admin_view(self.aprobar_una_orden), name='ordenes_ordenborrador_aprobar'),
            path('<int:orden_id>/eliminar/', self.admin_site.admin_view(self.eliminar_una_orden), name='ordenes_ordenborrador_eliminar'),
        ]
        return custom_urls + urls

    def aprobar_una_orden(self, request, orden_id):
        from django.shortcuts import redirect
        orden = self.get_object(request, str(orden_id))
        if orden:
            orden.estado = 'PENDIENTE'
            orden.save()
            self.message_user(request, f"Orden {orden.numero_orden} aprobada y enviada a la App.", messages.SUCCESS)
        return redirect('admin:ordenes_ordenborrador_changelist')

    def eliminar_una_orden(self, request, orden_id):
        from django.shortcuts import redirect
        orden = self.get_object(request, str(orden_id))
        if orden:
            numero = orden.numero_orden
            orden.delete()
            self.message_user(request, f"Orden {numero} eliminada permanentemente.", messages.WARNING)
        return redirect('admin:ordenes_ordenborrador_changelist')

    def botones_accion(self, obj):
        from django.urls import reverse
        return format_html(
            '<a style="background-color: #28a745; color: white; padding: 6px 12px; border-radius: 4px; font-weight: bold; margin-right: 5px; text-decoration: none; display: inline-block;" href="{}">✅Aprobar</a>'
            '<a style="background-color: #dc3545; color: white; padding: 6px 12px; border-radius: 4px; font-weight: bold; text-decoration: none; display: inline-block;" onclick="return confirm(\'¿Estás seguro de ELIMINAR la orden {}?\');" href="{}">🗑️ Borrar</a>',
            reverse('admin:ordenes_ordenborrador_aprobar', args=[obj.pk]),
            obj.numero_orden,
            reverse('admin:ordenes_ordenborrador_eliminar', args=[obj.pk]),
        )
    botones_accion.short_description = "Acciones Rápidas"
    botones_accion.allow_tags = True

    def ver_equipo(self, obj):
        if obj.equipo:
            return format_html("<b>{}</b><br><small>{}</small>", obj.equipo, obj.descripcion_equipo)
        return "-"
    ver_equipo.short_description = "Equipo"
    
    def ver_fechas(self, obj):
        if obj.inicio_programado:
            return obj.inicio_programado.strftime('%d/%m/%Y')
        return "-"
    ver_fechas.short_description = "Inicio"

# -----------------------
# PANTALLA 2: PENDIENTES
# -----------------------
@admin.register(OrdenPendiente)
class OrdenPendienteAdmin(admin.ModelAdmin):
    list_display = ('numero_orden', 'descripcion', 'codigo_trabajador', 'prioridad', 'ver_equipo_simple')
    search_fields = ('numero_orden', 'codigo_trabajador')
    inlines = [ActividadInline, EvidenciaInline]
    
    def get_queryset(self, request):
        return super().get_queryset(request).filter(estado='PENDIENTE')
        
    def has_add_permission(self, request):
        return False

    def ver_equipo_simple(self, obj):
        if obj.equipo:
             return f"{obj.equipo} - {obj.descripcion_equipo}"
        return "-"
    ver_equipo_simple.short_description = "Equipo"

# ----------------------
# PANTALLA 3: HISTORIAL
# ----------------------

class ActividadHistorialInline(admin.StackedInline):
    model = Actividad
    fk_name = 'orden'
    extra = 0
    can_delete = False
    
    verbose_name = "Operación Finalizada"
    verbose_name_plural = "Operaciones (Actividades)"
    
    readonly_fields = ('codigo_operacion', 'descripcion', 'nombre_ejecutor', 'fecha_inicio_real', 'fecha_fin_real', 'tiempo_legible', 'tiempo_pausas_legible')
    fields = ('codigo_operacion', 'descripcion', 'nombre_ejecutor', 'fecha_inicio_real', 'fecha_fin_real', 'tiempo_legible', 'tiempo_pausas_legible')
    
    def tiempo_legible(self, obj):
        if not obj.tiempo_real_acumulado:
            return "00:00:00"
        try:
            if hasattr(obj.tiempo_real_acumulado, 'total_seconds'):
                ts = int(obj.tiempo_real_acumulado.total_seconds())
            else:
                ts = int(obj.tiempo_real_acumulado) // 1000000
                
            horas, rem = divmod(ts, 3600)
            minutos, segundos = divmod(rem, 60)
            return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
        except Exception:
            return str(obj.tiempo_real_acumulado)
    tiempo_legible.short_description = "Tiempo Activo Real"

    def tiempo_pausas_legible(self, obj):
        if not obj.tiempo_pausas:
            return "00:00:00"
        try:
            if hasattr(obj.tiempo_pausas, 'total_seconds'):
                ts = int(obj.tiempo_pausas.total_seconds())
            else:
                ts = int(obj.tiempo_pausas) // 1000000
                
            horas, rem = divmod(ts, 3600)
            minutos, segundos = divmod(rem, 60)
            return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
        except Exception:
            return str(obj.tiempo_pausas)
    tiempo_pausas_legible.short_description = "Tiempo Total en Pausa"

    def has_add_permission(self, request, obj): 
        return False

@admin.register(OrdenHistorial)
class OrdenHistorialAdmin(admin.ModelAdmin):
    list_display = ('numero_orden', 'descripcion', 'codigo_trabajador', 'equipo', 'estado')
    search_fields = ('numero_orden', 'codigo_trabajador', 'equipo')
    list_filter = ('fin_programado',) 
    
    inlines = [ActividadHistorialInline, EvidenciaInline]

    @admin.action(description="🗑️ Eliminar órdenes seleccionadas")
    def eliminar_ordenes_seleccionadas(self, request, queryset): 
        cantidad = queryset.count()
        queryset.delete()
        self.message_user(request, f"Se eliminaron {cantidad} órdenes correctamente.") 

    @admin.action(description="📥 Exportar datos (Excel .xlsx)")
    def exportar_sap(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Reporte_Cierre.xlsx"'
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cierre de Órdenes"

        ws.sheet_view.showGridLines = False

        header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        borde_delgado = Border(
            left=Side(style='thin', color='000000'), 
            right=Side(style='thin', color='000000'), 
            top=Side(style='thin', color='000000'), 
            bottom=Side(style='thin', color='000000')
        )

        headers = [
            'Orden', 'Operación', 'Txt.brv.oper.', 'Equipo', 
            'Ubicación Téc.', 'Inicio Prog.', 'Fin Prog.',
            'Ejecutor Real', 'Inicio Real', 'Fin Real', 
            'Tiempo Activo (HH:MM:SS)', 'Tiempo en Pausa (HH:MM:SS)'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = borde_delgado

        def formatear_tiempo(valor):
            if not valor: return "00:00:00"
            try:
                if hasattr(valor, 'total_seconds'):
                    ts = int(valor.total_seconds())
                elif str(valor).isdigit():
                    ts = int(valor) // 1000000
                else:
                    return str(valor)
                horas, rem = divmod(ts, 3600)
                minutos, segundos = divmod(rem, 60)
                return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
            except Exception:
                return str(valor)

        for orden in queryset:
            actividades = Actividad.objects.filter(orden=orden)
            for act in actividades:
                t_activo = formatear_tiempo(act.tiempo_real_acumulado)
                t_pausa = formatear_tiempo(act.tiempo_pausas)

                inicio_real = act.fecha_inicio_real.strftime('%d/%m/%Y %H:%M') if act.fecha_inicio_real else '-'
                fin_real = act.fecha_fin_real.strftime('%d/%m/%Y %H:%M') if act.fecha_fin_real else '-'
                inicio_prog = orden.inicio_programado.strftime('%d/%m/%Y') if orden.inicio_programado else '-'
                fin_prog = orden.fin_programado.strftime('%d/%m/%Y') if orden.fin_programado else '-'

                fila = [
                    orden.numero_orden, act.codigo_operacion, act.descripcion,
                    orden.equipo, orden.ubicacion, inicio_prog, fin_prog,
                    act.nombre_ejecutor or '-', inicio_real, fin_real,
                    t_activo, t_pausa
                ]
                ws.append(fila)

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                cell.border = borde_delgado
                
                if cell.column >= 6:
                    cell.alignment = center_alignment
                
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
                    
            ajuste = (max_length + 2)
            ws.column_dimensions[column].width = ajuste

        wb.save(response)
        return response

    actions = ['exportar_sap', 'eliminar_ordenes_seleccionadas']

    def get_readonly_fields(self, request, obj=None):
        return [f.name for f in self.model._meta.fields]

    def has_add_permission(self, request): return False 
    def has_delete_permission(self, request, obj=None): return False 

    def get_queryset(self, request):
        return super().get_queryset(request).filter(estado='FINALIZADA')

    fields = (
        'numero_orden', 'descripcion', 'equipo', 'descripcion_equipo', 
        'ubicacion', 'ubicacion_tecnica', 'inicio_programado', 'fin_programado', 
        'prioridad', 'codigo_trabajador', 'supervisor', 'estado'
    )

# Limpieza
try:
    admin.site.unregister(User)
    admin.site.unregister(Group)
except: pass

@admin.register(User)
class CustomUserAdmin(UserAdmin):
    list_display = ('username', 'first_name', 'last_name', 'is_staff')